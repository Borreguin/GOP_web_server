"""
Created by rsanchez on 16/07/2018 - last version 14/08/2018
Este proyecto ha sido desarrollado en la Gerencia de Operaciones de CENACE
Mateo633
"""

import pandas as pd
import datetime as dt
import openpyxl as pyxl
from my_lib.GOP_connection import GOPserver as op
import sqlalchemy as sql_a

gop_svr = op.GOPserver()

""" Variables por defecto: """
layout_path_file = "S:/Aplicaciones Procesos/Aplicaciones SIVO/DATOSPROG_PLANTILLA.xlsx"
path_result_file = "P:/Despacho/DP_yyyymmdd.xlsx"
despacho_path_file = "M:/Despacho Prog/Pred_{0}.xlsx"
redespacho_path_file = "M:/Redespachos/A Ejecutarse/R{0}_{1}.xls"
# mem_file_path_file =  "M:/Genemp/Año2018/0718/25/DESPACHOPROG_2018-07-25.xlsx"
prog_file_path_file = "M:/Genemp/AñoYYYY/mmyy/dd/DESPACHOPROG_YYYY-mm-dd.xlsx"

id_columns = ['Empresa', 'UNegocio', 'Central', 'GrupoGeneracion', 'Unidad']
desp_columns = ['EsRedespacho', 'NumRedespacho', 'HoraVigencia']
data_columns = ['Fecha', 'Hora', 'MV', 'Precio']
table_columns = id_columns + desp_columns + data_columns

etiquetas_exportacion = ['ECUACOLO230',	'ECUACOLO138',	'ECUAPERU230']


def run_process_for(dt_date):
    msgs = "[Empezar]: \t\t  Empezando el proceso de {0}".format(dt_date)
    # run_tie_macro(dt_date)

    """ Variables de proceso """
    fecha_n = dt_date
    fecha_n_menos_1 = fecha_n - dt.timedelta(days=1)
    fecha_n_mas_1 = fecha_n + dt.timedelta(days=1)

    """ dia: n-1 """
    print("-->Iniciando: " + str(fecha_n_menos_1))
    # 1. Busca si existió un redespacho para el día n-1, caso contrario, busca el despacho:
    df_24_horas_n_menos_1, msg_i = obtener_hora_24_de_ultimo_despacho(fecha_n_menos_1)
    msgs += "\n\n--> Leyendo hora 24h: ({0})".format(fecha_n_menos_1)
    msgs += msg_i

    """ día: n """
    # 2.1 Despacho programado normal
    print("\n--> Iniciando: " + str(fecha_n))
    print("Leyendo despachos o redespachos")
    msgs += "\n\n--> Procesando despacho/redespacho de: " + str(fecha_n)
    d_path_file = despacho_path_file.format(dt_date.strftime('%y%m%d'))
    code_list, thermo_list, df_info_total, msg_n = \
        read_despacho(fecha_n, d_path_file)
    msgs += msg_n

    # Incluir la hora 24 del último despacho (n-1)
    df_info_total = pd.concat([df_24_horas_n_menos_1, df_info_total])
    df_info_total.sort_index(inplace=True)
    df_info_total = df_info_total.fillna(0)

    # La hora 24 del día n
    msg_hora_24h = "\n[Hora 24h]: \t \t Incluyendo hora 24h del despacho 0  de la fecha " + str(fecha_n)
    df_24_horas_n = df_info_total.iloc[[-1]]


    # incluir tipo de despacho:
    df_info_total.loc[:, "EsRedespacho"] = False
    df_info_total.loc[:, "NumRedespacho"] = 0
    df_info_total.loc[:, "HoraVigencia"] = "00:00:00"
    # tipo_despacho = dict(EsRedespacho=False, NumRedespacho=0, HoraVigencia="00:00:00")

    df_desp_to_save = df_info_total.copy()  # info despacho to save in excel

    # Excluir la hora 24 del siguiente despacho (n+1)
    df_info_total = df_info_total.iloc[:-1]

    msgs += run_despacho(fecha_n, df_info_total, code_list, thermo_list)

    # 2.2. Busca redespachos para el día n
    df_redesp_to_save = pd.DataFrame()
    for n_redespacho in range(1, 10):
        code_list, thermo_list, df_info_total, hora_vigente, msg_n = \
            read_redespacho(fecha_n, n_redespacho)
        if df_info_total.empty:
            if n_redespacho == 1:
                msgs += "\n\n[Leer redespachos]: \t No hay redespachos para la fecha " + str(fecha_n)
            break
        if hora_vigente == "No hora vigente identificada":
            msgs += msg_n + "\n error: redespacho no subido: No hora vigente identificada"
            break

        else:
            # La hora 24 del día n (redespacho)
            msg_hora_24h = "\n[Hora 24h]: \t \t Incluyendo hora 24h del redespacho {0} de la fecha {1}" \
                .format(n_redespacho, fecha_n)

            # seleccionar la hora 24h
            df_24_horas_n = df_info_total.iloc[[-1]]


            df_redespacho_n = df_info_total.copy()

            # Incluye la hora 00 del último despacho (n-1)
            df_redespacho_n = pd.concat([df_24_horas_n_menos_1, df_redespacho_n])

            # Existio redespacho, modificando el tipo:
            # tipo_despacho = dict(EsRedespacho=True, NumRedespacho=n_redespacho, HoraVigencia=str(hora_vigente))
            df_redespacho_n.loc[:, "EsRedespacho"] = True
            df_redespacho_n.loc[:, "NumRedespacho"] = n_redespacho
            df_redespacho_n.loc[:, "HoraVigencia"] = str(hora_vigente)
            df_redespacho_n.fillna(0, inplace=True)


            # Para guardar el archivo de excel:
            new_index = list(df_redespacho_n.index)
            new_index[-1] = df_redespacho_n.index[-1] - pd.Timedelta("1 s")
            df_aux = df_redespacho_n.copy()
            df_aux.index = new_index
            df_redesp_to_save = pd.concat([df_redesp_to_save, df_aux])

            # Excluye la hora 24 del día n
            df_redespacho_n = df_redespacho_n.iloc[:-1]

            msgs += msg_n
            msgs += run_despacho(fecha_n, df_redespacho_n, code_list, thermo_list)

    # df_to_save y tipo_despacho contiene la última información

    new_index = list(df_desp_to_save.index)
    new_index[-1] = df_desp_to_save.index[-1] - pd.Timedelta("1 s")  # 23:59
    df_desp_to_save.index = new_index

    if df_redesp_to_save.empty:
        df_to_save = df_desp_to_save[1:]
    else:
        # incluye la hora 24h del dia n-1, es decir las 0:00 del dia n
        # df_redesp_to_save = pd.concat([df_24_horas_n_menos_1, df_redesp_to_save])
        # df_redesp_to_save.index = new_index
        df_to_save = pd.concat([df_desp_to_save[1:], df_redesp_to_save[1:]])
        df_to_save.fillna(0, inplace=True)


    msgs += ("\n\n--> Guardar Excel de la fecha " + str(fecha_n))
    # excluye la hora 00 del día n
    # df_to_save = df_to_save.iloc[1:]
    msgs += run_despacho(fecha_n, df_to_save, code_list, thermo_list, saveSIVO=False, saveExcel='MEM')

    """ día: n+1 """
    msgs += "\n\n--> Procesando Despacho de: " + str(fecha_n_mas_1)
    msgs += msg_hora_24h
    # 3.1. Obtener información del despacho normal:
    d_path_file = despacho_path_file.format(fecha_n_mas_1.strftime('%y%m%d'))
    code_list, thermo_list, df_info_total, msg_i = read_despacho(fecha_n_mas_1, d_path_file)
    msgs += msg_i
    if not df_info_total.empty:
        # Excluye la hora 24 del día n+1 (00:00 del n+2)
        df_despacho_n_mas_1 = df_info_total.iloc[:-1]

        # Incluye la hora 00:00 del día n (sea por despacho o por redespacho)
        df_despacho_n_mas_1 = pd.concat([df_24_horas_n, df_despacho_n_mas_1])
        df_despacho_n_mas_1.sort_index(inplace=True)

        # Procesar información del despacho programado:
        # desp_normal = dict(EsRedespacho=False, NumRedespacho=0, HoraVigencia='00:00:00')
        df_despacho_n_mas_1.loc[:, "EsRedespacho"] = False
        df_despacho_n_mas_1.loc[:, "NumRedespacho"] = 0
        df_despacho_n_mas_1.loc[:, "HoraVigencia"] = "00:00:00"

        msgs += run_despacho(fecha_n_mas_1,
                             df_info_total=df_despacho_n_mas_1,
                             code_list=code_list,
                             thermo_list=thermo_list, detalle=True)

    if "error" in msgs or "No existe el código" in msgs:
        title = "[SIVO][ALERTA-REVISAR] Despacho programado subido automáticamente: [Fecha: {0}]".format(dt_date)
    else:
        title = "[SIVO]Despacho programado subido automáticamente: [Fecha: {0}]".format(dt_date)
    send_mail(msg_to_send=msgs, subject=title)
    # print(msgs.encode('ascii', 'ignore'))

    return title, msgs


def obtener_hora_24_de_ultimo_despacho(fecha_t):
    msg_i = str()
    msg_24h = str()
    """ Busando la hora 24 en redespachos: """
    df_24_horas = pd.DataFrame()
    for n_redespacho in range(1, 10):
        code_list, thermo_list, df_info_total, hora_vigente, msg_n = \
            read_redespacho(fecha_t, n_redespacho)
        if df_info_total.empty:
            if n_redespacho == 1:
                msg_i += "\n[Leer Redespachos]: \tNo hay redespacho para la fecha " + str(fecha_t)
            break
        else:
            df_24_horas = df_info_total.iloc[[-1]]
            msg_i += msg_n
            msg_24h = "\n[Hora 24h]: \t \t Incluyendo hora 24h del redespacho {0} de la fecha {1}" \
                .format(n_redespacho, fecha_t)

    """ Si no hubo redespachos, entonces en despachos """
    if df_24_horas.empty:
        d_path_file = despacho_path_file.format(fecha_t.strftime('%y%m%d'))
        code_list, thermo_list, df_info_total, msg_n = read_despacho(fecha_t, d_path_file)
        df_24_horas = df_info_total.iloc[[-1]]
        msg_24h = "\n[Hora 24h]: \t \t Incluyendo hora 24h del despacho 0 de la fecha " + str(fecha_t)
        msg_i += msg_n

    return df_24_horas, msg_i + msg_24h


def read_despacho(dt_date, path_file):
    """  Leyendo el archivo de despacho """
    # Obteniendo los códigos de las centrales de generación
    code_list, thermo_list = get_codigos(path_file)
    if len(code_list) == 0:
        return list(), list(), pd.DataFrame(), "\n[{0}]: El archivo no existe".format(path_file)

    df_info_total = get_info_despacho(dt_date, path_file)
    msg_i = "\n[Leer Despacho]: \t [{0}]: Archivo de despacho leído.".format(path_file)
    return code_list, thermo_list, df_info_total, msg_i


def read_redespacho(dt_date, n_redespacho):
    msg_i = str()
    # R1_2018-07-30
    """ Leyendo el archivo de redespacho"""
    r_path_file = redespacho_path_file.format(n_redespacho, dt_date.strftime('%Y-%m-%d'))
    d_path_file = despacho_path_file.format(dt_date.strftime('%y%m%d'))
    print("Archivos a leer: \n \t{0}  \n \t{1}".format(r_path_file, d_path_file))
    hora_vigente = "00:00:00"
    df_info_total = read_excel_file(r_path_file, skiprows=10, sheet_name="REDESPACHO")

    if df_info_total.empty:
        msg_i += "[Leer redespacho] [{0}]: No existe redespacho".format(r_path_file)

    else:
        df_info_total = df_info_total.drop([0, 1])[:24]
        valid_columns = [x for x in df_info_total.columns if 'Unnamed' not in x]
        valid_columns = [x for x in valid_columns if 'HORA.1' not in x]
        df_info_total = df_info_total[valid_columns]
        df_info_total.index = pd.date_range(dt_date, dt_date + pd.Timedelta('24 H'), freq="60T", closed='right')
        df_info_total.fillna(0, inplace=True)
        df_info_total.drop(["HORA"], axis=1, inplace=True)
        hora_vigente = get_hora_vigente(r_path_file)
        msg_i += "\n[Leer redespacho]: \t [{0}]: Archivo de redespacho leído.".format(r_path_file)

    # Obteniendo los códigos de las centrales de generación
    code_list, thermo_list = get_codigos(d_path_file)
    if len(code_list) == 0:
        return list(), list(), pd.DataFrame(), hora_vigente, "\n[{0}]: El archivo no existe".format(d_path_file)

    return code_list, thermo_list, df_info_total, hora_vigente, msg_i


def convert_str_to_time(str_time, ls_format=None):
    if ls_format is None:
        ls_format = ["%H:%M:%S", "%H:%M", "%Y-%m-%d %H:%M:%S"]
    dt_resp = None
    for fmt in ls_format:
        try:
            dt_resp = dt.datetime.strptime(str_time, fmt)
            break
        except ValueError:
            pass
    return dt_resp



def get_hora_vigente(rcho_path_file):
    df_info = read_excel_file(rcho_path_file, skiprows=8, sheet_name="REDESPACHO")
    hora_vigente = "00:00"
    for ix in list(range(2, 6)):
        hora_vigente = df_info.columns[ix]
        if isinstance(hora_vigente, dt.time):
            break
        else:
            hora_vigente = convert_str_to_time(hora_vigente)
            if hora_vigente is not None:
                hora_vigente = hora_vigente.strftime("%H:%M:%S")
                break

    if "Unnamed" not in str(hora_vigente):
        return hora_vigente
    else:
        return "No hora vigente identificada"


def run_despacho(dt_date, df_info_total, code_list, thermo_list, saveSIVO=True,
                 saveExcel='Default', detalle=False):
    msg_i = str()
    # Filtrando informaión del despacho
    try:
        other_information = ['DEMANDA', 'PRECIO', 'RSF']
        df_despacho = df_info_total.drop(other_information, axis=1)
        df_other_info = df_info_total[other_information]

    except Exception as e:
        other_information = ['DEMANDA', 'PRECIO']
        df_despacho = df_info_total.drop(other_information, axis=1)
        df_other_info = df_info_total[other_information]

    # Obteniedo la tabla de mapeo:
    df_map, etiquetas_adicionales = get_mapping_info(layout_path_file)

    """ A REPORTAR: Si la central se encuentra inhabilitada 
        pero existe dentro de la lista de códigos de planeamiento 
    """
    list_report = set(code_list).intersection(set(df_map[df_map["HABILITADOR"] == 0]["ETIQUETAS"]))

    """ Lista de centrales operativas: 
        la lista de las centrales de planeamiento 
        y las habilitadas en la lista de mapeo
    """
    list_to_work = set(code_list).intersection(set(df_map[df_map["HABILITADOR"] == 1]["ETIQUETAS"]))
    list_to_work = list(list_to_work.union(df_despacho.columns)) + etiquetas_adicionales
    list_to_work = list(set(list_to_work).difference(desp_columns))
    list_to_work.sort()

    # Creando los valores de fecha y hora:
    fecha_values = [x.strftime('%Y-%m-%d') for x in df_despacho.index]
    hora_values = [x.strftime('%H:%M') for x in df_despacho.index]

    # Iniciando el dataframe de respuesta:
    df_result = pd.DataFrame(columns=table_columns)

    # Trabajando con la lista de centrales operativas
    for code in list_to_work:
        # print(code)
        df_i = pd.DataFrame(columns=table_columns, index=df_despacho.index)

        # Usando información de la tabla de mapeo
        try:
            mask = (df_map["ETIQUETAS"] == code)
            df_filter = df_map[mask][id_columns]

            for col_i in id_columns:
                df_i[col_i] = df_filter[col_i].iloc[0]

        except Exception as e:
            msg_i += "\n[ALERTA CÓDIGOS] revisar lo siguiente: \n[{0}] No existe el código {1} en la tabla MAPA ".format(layout_path_file, code)
            print(msg_i)
            continue  # Se esquiva este código y se continua con el siguiente

        # LLenando valores del tipo de despacho
        for col_i in desp_columns:
            # df_i[col_i] = tipo_de_despacho[col_i]
            df_i[col_i] = df_info_total[col_i]

        # Llenando valores de la tabla:
        df_i['Fecha'] = fecha_values
        df_i['Hora'] = hora_values
        df_i['Precio'] = df_other_info["PRECIO"]

        # si la central fue despachada:
        if code in df_despacho.columns:
            df_i['MV'] = df_despacho[code]
        elif code in list(thermo_list):
            # caso contrario si es térmica y no fue despachada, encerar
            df_i['MV'] = 0
        elif code in etiquetas_adicionales:
            #  caso contrario si es un caso especial, encerar
            df_i['MV'] = 0

        df_i['NombreArchivo'] = "Carga_automática " + str(dt_date)

        df_result = pd.concat([df_result, df_i])

    if saveSIVO:
        msg_i += subir_a_SIVO(df_result)

    if saveExcel == 'Default':
        path_file = path_result_file.replace("yyyymmdd", dt_date.strftime('%Y%m%d'))
        msg_i += save_result_to_excel(df_result, path_file)
    elif saveExcel == 'MEM':
        path_file = prog_file_path_file.replace("YYYY", dt_date.strftime("%Y"))
        path_file = path_file.replace("mm", dt_date.strftime("%m"))
        path_file = path_file.replace("dd", dt_date.strftime("%d"))
        path_file = path_file.replace("yy", dt_date.strftime("%y"))
        df_result.rename(index=str, columns={"MV": "MW"}, inplace=True)
        order_list = ['Empresa', 'UNegocio', 'Central', 'GrupoGeneracion', 'Unidad', 'EsRedespacho', 'NumRedespacho',
                      'HoraVigencia', 'Fecha', 'Hora', 'MW', 'Precio']

        mask = df_map["ETIQUETAS"].isin(etiquetas_exportacion)
        grupo_generacion = list(df_map["GrupoGeneracion"][mask])
        mask = df_result["GrupoGeneracion"].isin(grupo_generacion)
        df_result.fillna(0, inplace=True)
        df_result.loc[mask, "MW"] = - df_result[mask]["MW"]
        msg_i += save_result_to_excel(df_result, path_file, sheet_name='DPL_DespachoProgramado', columns=order_list)


    msg_i += "\n[Check entidades]: \t Un total de {0} entidades fueron procesadas".format(len(code_list))
    msg_i += "\n[Fechas proces.]: \t Desde {0} hasta {1}".format(df_despacho.index[0], df_despacho.index[-1])
    msg_i += ("\n[Unidades deshabilitadas]: " + list_in_columns(list(list_report)))

    if "error" in msg_i or "Error" in msg_i:
        msg_i += "\n[Proceso]: \t\t Proceso finalizado con errores "
    else:
        msg_i += "\n[Proceso]: \t\t Proceso finalizado de manera exitosa "

    if detalle:
        msg_i += ("\n\nDetalle: " + list_in_columns(code_list))

    return msg_i


def list_in_columns(foolist):
    str_l = str()
    for a, b, c in zip(foolist[::3], foolist[1::3], foolist[2::3]):
        str_l += '\n{:<18}\t{:18}\t{:<}'.format(a, b, c)

    return str_l


def send_mail(msg_to_send, subject):
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    import smtplib

    # create message object instance
    msg = MIMEMultipart()

    saludo = "Estimados ingenieros, \nA continuación el resumen del " \
             "proceso automático de subida del despacho programado a SIVO: \n \n"

    message = saludo + msg_to_send

    # setup the parameters of the message
    password = "SAOGMAIL"
    recipients = ["ugie@cenace.org.ec", "rsanchez@cenace.org.ec"]
    msg['From'] = "aadopost.cenace@gmail.com"
    msg['To'] = ",".join(recipients)
    msg['Subject'] = subject

    # add in the message body
    msg.attach(MIMEText(message, 'plain'))

    # create server
    server = smtplib.SMTP('smtp.gmail.com: 587')

    server.starttls()

    # Login Credentials for sending the mail
    server.login(msg['From'], password)

    # send the message via the server.
    server.sendmail(msg['From'], msg['To'], msg.as_string())

    server.quit()

    print("Detalles enviados a: %s:" % (msg['To']))


def save_result_to_excel(df_result, path_file, **kwargs):
    try:
        df_result.to_excel(path_file, index=False, **kwargs)
        msg = "\n[Guardar Excel]: \t Archivo [{0}] guardado exitosamente".format(path_file)
    except Exception as e:
        new_path_file = path_file.replace('.xlsx', str(dt.datetime.now()).replace(":", "_") + '.xlsx')
        msg = "\n[Guardar Excel]: \t Error al guardar el archivo [{0}]. Guardando como: ".format(new_path_file)
        df_result.to_excel(new_path_file, index=False)

    return msg


def subir_a_SIVO(df_result):
    msg = str()
    try:
        df_result.to_sql(name="DPL_DespachoProgramado", if_exists='append', index=False, con=gop_svr.engine)
        msg += "\n[Subir a DB SIVO]: \t Despacho programado subido de manera existosa a SIVO"
    except Exception as exp:
        list_generation_group = list(set(df_result["GrupoGeneracion"]))
        list_generation_group.sort()

        err_i, exp = 0, None
        entities = set()
        for group in list_generation_group:
            try:
                mask = (df_result["GrupoGeneracion"] == group)
                df_aux = df_result[mask]
                df_aux.to_sql(name="DPL_DespachoProgramado", if_exists='append', index=False, con=gop_svr.engine)
                entities.add(group)

            except Exception as e:
                err_i += 1

        if err_i == len(list_generation_group):
            msg += "\n[Subir a DB SIVO]: \t Despacho programado ya existente"
            pass
        elif len(entities) > 0:
            msg += "\n[Subir a DB SIVO]: \t Las siguientes entidades han sido añadidas al despacho programado: \n" + str(entities)
        else:
            msg += "\n[Subir a DB SIVO]: \t " + str(e)
    return msg


def get_info_despacho(dt_date, path_file):
    df = read_excel_file(path_file, skiprows=8, sheet_name="DESPACHO")
    df = df.drop([0, 1])[:24]
    valid_columns = [x for x in df.columns if 'Unnamed' not in x]
    valid_columns = [x for x in valid_columns if 'HORA.1' not in x]
    df = df[valid_columns]
    df.index = pd.date_range(dt_date, dt_date + pd.Timedelta('24 H'), freq="60T", closed='right')
    df.fillna(0, inplace=True)
    return df


def get_codigos(path_file):
    df_t = read_excel_file(path_file, sheet_name="CÓDIGOS")
    if df_t.empty:
        return list(), set()
    df_hydro = df_t["Unidades Hidráulicas "].dropna()
    df_thermo = df_t["Unidades Térmicas"].dropna()
    df_no_conventional = df_t["Unidades Energía no Convencional"].dropna()
    code_list = list(df_hydro) + list(df_thermo) + list(df_no_conventional)
    exclude = ['Códigos de despacho', 'TEXTO']
    code_list = list(set([x for x in code_list if x not in exclude]))
    code_list.sort()
    total_list = set(list(df_thermo) + list(df_hydro) + list(df_no_conventional))
    return code_list, total_list


def get_mapping_info(l_path_file):
    df_map = pd.read_excel(l_path_file, sheet_name="MAPA")
    df_etiquetas_especiales = pd.read_excel(l_path_file, sheet_name="ADICIONALES")
    valid_columns = [x for x in df_map.columns if 'Unnamed' not in str(x)]
    valid_columns = [x for x in valid_columns if 'HORA.1' not in x]
    df_map = df_map[valid_columns]
    # df_map = df_map[df_map["HABILITADOR"] == 1]
    lst = df_etiquetas_especiales.T.values
    return df_map, list(lst[0])


def run_tie_macro(dt_date):
    # creando una hoja virtual (dataFrame)
    ini_time = pd.to_datetime(dt_date.date())
    end_time = ini_time + pd.Timedelta('23 H 30 m')
    time_range = pd.date_range(ini_time, end_time, freq="30T")
    time_range = list(time_range) + [time_range[0] + pd.Timedelta('23 H 59 m 59 s')]
    columns = ['LINEA', 'FECHA', 'HORA', 'TRANSACCION']
    df_result_138_kV = pd.DataFrame(index=time_range, columns=columns)
    df_result_230_kV = pd.DataFrame(index=time_range, columns=columns)
    df_result_138_kV = give_format(df_result_138_kV, 'PAN138TUL')
    df_result_230_kV = give_format(df_result_230_kV, 'JAM230POM')

    # leyendo el archivo de Resultado:
    path_file = "M:\Datos Predes\Resultado{0}.xls".format(dt_date.strftime("%d%m"))
    df = read_excel_file(path_file, skiprows=8)
    valid_columns = [x for x in df.columns if 'Unnamed' not in x]
    valid_columns = [x for x in valid_columns if 'HORA.1' not in x]
    df = df[valid_columns]

    # existe transaccion
    df_result_230_kV['TRANSACCION'] = hay_transaccion(df.iloc[:, 1])
    df_result_138_kV['TRANSACCION'] = hay_transaccion(df.iloc[:, 6])

    # save results:
    df_final = pd.concat([df_result_138_kV, df_result_230_kV])
    df_final = df_final[columns]
    path_to_save = "P:\Resultado\RE_{0}.xlsx".format(dt_date.strftime("%Y%m%d"))
    df_final.to_excel(path_to_save, index=False, sheet_name="Resultado")


# Auxiliary functions:
def read_excel_file(path_file, **kwargs):
    df = pd.DataFrame()
    try:
        df = pd.read_excel(path_file, **kwargs)
    except Exception as e:
        pass
    return df


def give_format(df, nombre_de_linea):
    df['LINEA'] = nombre_de_linea
    df['FECHA'] = [x.date() for x in df.index]
    df['HORA'] = [str(x.time()) for x in df.index]
    return df


def hay_transaccion(df):
    transaccion = list()
    transaccion_i = None
    for x in df.dropna():
        if "COLOMBIA" in x or "- - -" in x:
            transaccion_i = 'N'
        else:
            transaccion_i = 'S'

        # para t = xx:00
        transaccion.append(transaccion_i)
        # para t = xx:30
        transaccion.append(transaccion_i)

    # para t = 59:59
    transaccion.append(transaccion_i)
    return transaccion[1:]


def write_in_xls_file(path_file, sheet_name, df, place, **kwargs):
    # read file:
    srcfile = pyxl.load_workbook(path_file, **kwargs)  # to open the excel sheet and if it has macros
    sheetname = srcfile[sheet_name]  # get sheet_name from the file

    # sheetname['B2'] = str('write something')  # write something in B2 cell of the supplied sheet
    row_index = place['row']
    column_index = place['column']

    if isinstance(df, pd.Series):

        for row_i in range(len(df.index)):
            write_row = row_index + row_i
            sheetname.cell(row=write_row, column=column_index).value = df.iloc[row_i]
    else:

        for col_i, column in enumerate(df.columns):
            for row_i in range(len(df[column].index)):
                write_row = row_index + row_i
                write_column = column_index + col_i - 1
                sheetname.cell(row=write_row, column=write_column).value = df[column].iloc[row_i]

    srcfile.save(path_file.replace(".xlsx", "_new.xlsx"))


if __name__ == "__main__":
    today = True
    if today:
        dt_today = dt.datetime.today().date()
    else:
        dt_today = dt.datetime.strptime("2018-11-12", "%Y-%m-%d").date()
    print("Empezando proceso por:" + str(dt_today))
    run_process_for(dt_today)

""" 
dt_plus_1 = dt_date +  dt.timedelta(days=1)
sql_str = "DELETE FROM dbo.DPL_DespachoProgramado " \
          "WHERE Fecha ='{0}' " \
          "or (Fecha = '{1}' " \
          " and  Hora='00:00:00');".format(
            dt_date.strftime("%Y-%m-%d"),
            dt_plus_1.strftime("%Y-%m-%d")
            )
print(sql_str)
gop_svr.engine.execute(sql_str)
"""
